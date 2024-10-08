from flask import (
    Blueprint, jsonify, request, send_file, url_for, redirect, flash
)
import io
from .db import db
from .helper import admin_required
from .models import Registrant, RegistrantData, Parent
from sqlalchemy.orm import aliased
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tempfile import NamedTemporaryFile

bp = Blueprint('export', __name__, url_prefix='/export')


@bp.route('/data_pendaftar', methods=['POST'])
@admin_required
def data_pendaftar():
    selection_path = request.form['selection_path']
    program = request.form['program']
    rg = aliased(Registrant)
    rgd = aliased(RegistrantData)
    ft = aliased(Parent)
    mt = aliased(Parent)
    gd = aliased(Parent)
    qry = db.session.query(rg).\
        outerjoin(rgd, rg.registrant_data_id == rgd.id).\
        outerjoin(ft, rg.father_id == ft.id).\
        outerjoin(mt, rg.mother_id == mt.id).\
        outerjoin(gd, rg.guardian_id == gd.id)
        
    if selection_path is None or program is None:
        flash('Pilih program dan jalur pendaftaran', 'error')
        return redirect(url_for('admin.lihat_pendaftar'))
    
    if selection_path != 'Semua':
        qry = qry.filter(rg.selection_path == selection_path)
    if program != 'Semua':
       qry = qry.filter(rg.program == program)
    
    result = qry.filter(rg.deleted == False).with_entities(
        # urutan di excel tergantung urutan di sini
            rg.reg_id.label('reg_id'),
            rg.name.label('name'),
            rg.gender.label('gender'),
            rg.prev_school.label('prev_school'),
            rg.nisn.label('nisn'),
            rg.cp.label('cp'),
            rg.program.label('program'),
            rg.selection_path.label('selection_path'),
            rg.entry_year.label('entry_year'),
            rg.gelombang.label('gelombang'),
            rg.initial_cost.label('initial_cost'),
            rg.monthly_cost.label('monthly_cost'),
            rg.land_donation.label('land_donation'),
            rg.qurban.label('qurban'),
            rgd.nik.label('nik'),
            rgd.nkk.label('nkk'),
            rgd.nak.label('nak'),
            rgd.birth_place.label('birth_place'),
            rgd.birth_date.label('birth_date'),
            rgd.birth_order.label('birth_order'),
            rgd.siblings_count.label('siblings_count'),
            rgd.street.label('street'),
            rgd.rt.label('rt'),
            rgd.rw.label('rw'),
            rgd.village.label('village'),
            rgd.district.label('district'),
            rgd.city.label('city'),
            rgd.province.label('province'),
            rgd.country.label('country'),
            rgd.postal_code.label('postal_code'),
            rgd.stay_with.label('stay_with'),
            rgd.parent_status.label('parent_status'),
            rgd.nationality.label('nationality'),
            rgd.religion.label('religion'),
            rgd.height.label('height'),
            rgd.weight.label('weight'),
            rgd.head_size.label('head_size'),
            rgd.hobbies.label('hobbies'),
            rgd.hospital_sheets.label('hospital_sheets'),
            rgd.physical_abnormalities.label('physical_abnormalities'), #AN
            ft.name.label('ft_name'),
            ft.nik.label('ft_nik'),
            ft.status.label('ft_status'),
            ft.birth_place.label('ft_birth_place'),
            ft.birth_date.label('ft_birth_date'),
            ft.street.label('ft_street'),
            ft.rt.label('ft_rt'),
            ft.rw.label('ft_rw'),
            ft.village.label('ft_village'),
            ft.district.label('ft_district'),
            ft.city.label('ft_city'),
            ft.province.label('ft_province'),
            ft.country.label('ft_country'),
            ft.postal_code.label('ft_postal_code'),
            ft.contact.label('ft_contact'),
            ft.relation.label('ft_relation'),
            ft.nationality.label('ft_nationality'),
            ft.religion.label('ft_religion'),
            ft.education_level.label('ft_education_level'),
            ft.job.label('ft_job'),
            ft.position.label('ft_position'),
            ft.company.label('ft_company'),
            ft.income.label('ft_income'),
            ft.burden_count.label('ft_burden_count'), #BL
            mt.name.label('mt_name'),
            mt.nik.label('mt_nik'),
            mt.status.label('mt_status'),
            mt.birth_place.label('mt_birth_place'),
            mt.birth_date.label('mt_birth_date'),
            mt.street.label('mt_street'),
            mt.rt.label('mt_rt'),
            mt.rw.label('mt_rw'),
            mt.village.label('mt_village'),
            mt.district.label('mt_district'),
            mt.city.label('mt_city'),
            mt.province.label('mt_province'),
            mt.country.label('mt_country'),
            mt.postal_code.label('mt_postal_code'),
            mt.contact.label('mt_contact'),
            mt.relation.label('mt_relation'),
            mt.nationality.label('mt_nationality'),
            mt.religion.label('mt_religion'),
            mt.education_level.label('mt_education_level'),
            mt.job.label('mt_job'),
            mt.position.label('mt_position'),
            mt.company.label('mt_company'),
            mt.income.label('mt_income'),
            mt.burden_count.label('mt_burden_count'), #CJ
            gd.name.label('gd_name'),
            gd.nik.label('gd_nik'),
            gd.status.label('gd_status'),
            gd.birth_place.label('gd_birth_place'),
            gd.birth_date.label('gd_birth_date'),
            gd.street.label('gd_street'),
            gd.rt.label('gd_rt'),
            gd.rw.label('gd_rw'),
            gd.village.label('gd_village'),
            gd.district.label('gd_district'),
            gd.city.label('gd_city'),
            gd.province.label('gd_province'),
            gd.country.label('gd_country'),
            gd.postal_code.label('gd_postal_code'),
            gd.contact.label('gd_contact'),
            gd.relation.label('gd_relation'),
            gd.nationality.label('gd_nationality'),
            gd.religion.label('gd_religion'),
            gd.education_level.label('gd_education_level'),
            gd.job.label('gd_job'),
            gd.position.label('gd_position'),
            gd.company.label('gd_company'),
            gd.income.label('gd_income'),
            gd.burden_count.label('gd_burden_count'), #DH
        ).order_by(rg.id).all()
    
    file_content = buat_konten(result)
    return send_file(io.BytesIO(file_content), as_attachment=True, download_name='data_pendaftar.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #return jsonify(data)
    
def buat_konten(data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Pendaftar'
    ws['A4'].value = 'Data Pendaftar'
    ws.merge_cells('A4:AN4')
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AO4'].value = 'Data Ayah'
    ws.merge_cells('AO4:BL4')
    ws['AO4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['BM4'].value = 'Data Ibu'
    ws.merge_cells('BM4:CJ4')
    ws['BM4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['CK4'].value = 'Data Wali'
    ws.merge_cells('CK4:DI4')
    ws['CK4'].alignment = Alignment(horizontal="center", vertical="center")
    
    labels = [
        ('reg_id', 'Nomor Pendaftaran'),
        ('name', 'Nama'),
        ('gender', 'P/L'),
        ('prev_school', 'Sekolah Asal'),
        ('nisn', 'NISN'),
        ('cp', 'Kontak Utama/Pendaftar'),
        ('program', 'Program'),
        ('selection_path', 'Jalur Seleksi Pendaftaran'),
        ('entry_year', 'Tahun Masuk'),
        ('gelombang', 'Gelombang'),
        ('initial_cost', 'Uang Pangkal'),
        ('monthly_cost', 'SPP'),
        ('land_donation', 'Wakaf Tanah'),
        ('qurban', 'Tahun Qurban'),
        ('nik', 'No. Induk Kependudukan'),
        ('nkk', 'No. Kartu Keluarga'),
        ('nak', 'No. Akte Pendaftaran'),
        ('birth_place', 'Tempat Lahir'),
        ('birth_date', 'Tgl. Lahir'),
        ('birth_order', 'Anak Ke'),
        ('siblings_count', 'Dari .. Bersaudara'),
        ('street', 'Dusun/Jalan'),
        ('rt', 'rt'),
        ('rw', 'rw'),
        ('village', 'Desa/Kelurahan'),
        ('district', 'Kecamatan'),
        ('city', 'Kota/Kabupaten'),
        ('province', 'Provinsi'),
        ('country', 'Negara'),
        ('postal_code', 'Kode Post'),
        ('stay_with', 'Tinggal Bersama'),
        ('parent_status', 'Status Orang Tua'),
        ('nationality', 'Kewarganegaraan'),
        ('religion', 'Agama'),
        ('height', 'Tinggi'),
        ('weight', 'Berat'),
        ('head_size', 'Lingkar Kepala'),
        ('hobbies', 'Hobi'),
        ('hospital_sheets', 'Catatan Kesehatan'),
        ('physical_abnormalities', 'Kelainan Fisik'), #AN
        ('ft_name', 'Nama Ayah'),
        ('ft_nik', 'NIK Ayah'),
        ('ft_status', 'Status Ayah'),
        ('ft_birth_place', 'Tempat Lahir'),
        ('ft_birth_date', 'Tanggal Lahir'),
        ('ft_street', 'Dusun/Jalan'),
        ('ft_rt', 'RT'),
        ('ft_rw', 'RW'),
        ('ft_village', 'Desa/Kelurahan'),
        ('ft_district', 'Kecamatan'),
        ('ft_city', 'Kota/Kabupaten'),
        ('ft_province', 'Provinsi'),
        ('ft_country', 'Negara'),
        ('ft_postal_code', 'Kode Pos'),
        ('ft_contact', 'Kontak Ayah'),
        ('ft_relation', 'Hub. dengan Pendaftar'),
        ('ft_nationality', 'Kewarganegaraan'),
        ('ft_religion', 'Agama'),
        ('ft_education_level', 'Pendidikan Terakhir'),
        ('ft_job', 'Pekerjaan'),
        ('ft_position', 'Jabatan/Posisi'),
        ('ft_company', 'Tempat Kerja/Perusahaan'),
        ('ft_income', 'Penghasilan Bulanan'),
        ('ft_burden_count', 'Jumlah Tanggungan'), #BL
        ('mt_name', 'Nama Ibu'),
        ('mt_nik', 'NIK Ibu'),
        ('mt_status', 'Status Ibu'),
        ('mt_birth_place', 'Tempat Lahir'),
        ('mt_birth_date', 'Tgl. Lahir'),
        ('mt_street', 'Dusun/Jalan'),
        ('mt_rt', 'RT'),
        ('mt_rw', 'RW'),
        ('mt_village', 'Desa/Kelurahan'),
        ('mt_district', 'Kecamatan'),
        ('mt_city', 'Kota/Kabupaten'),
        ('mt_province', 'Provinsi'),
        ('mt_country', 'Negara'),
        ('mt_postal_code', 'Kode Pos'),
        ('mt_contact', 'Kontak Ibu'),
        ('mt_relation', 'Hub. dengan Pendaftar'),
        ('mt_nationality', 'Kewarganegaraan'),
        ('mt_religion', 'Agama'),
        ('mt_education_level', 'Pendidikan Terakhir'),
        ('mt_job', 'Pekerjaan'),
        ('mt_position', 'Jabatan/Posisi'),
        ('mt_company', 'Tempat Kerja/Perusahaan'),
        ('mt_income', 'Penghasilan Bulanan'),
        ('mt_burden_count', 'Jumlah Tanggungan'), #CU
        ('gd_name', 'Nama Wali'),
        ('gd_nik', 'NIK Wali'),
        ('gd_status', 'Status Wali'),
        ('gd_birth_place', 'Tempat Lahir'),
        ('gd_birth_date', 'Tgl. Lahir'),
        ('gd_street', 'Dusun/Jalan'),
        ('gd_rt', 'RT'),
        ('gd_rw', 'RW'),
        ('gd_village', 'Desa/Kelurahan'),
        ('gd_district', 'Kecamatan'),
        ('gd_city', 'Kota/Kabupaten'),
        ('gd_province', 'Provinsi'),
        ('gd_country', 'Negara'),
        ('gd_postal_code', 'Kode Pos'),
        ('gd_contact', 'Kontak Wali'),
        ('gd_relation', 'Hub. dengan Pendaftar'),
        ('gd_nationality', 'Kewarganegaraan'),
        ('gd_religion', 'Agama'),
        ('gd_education_level', 'Pendidikan Terakhir'),
        ('gd_job', 'Pekerjaan'),
        ('gd_position', 'Jabatan/Posisi'),
        ('gd_company', 'Tempat Kerja/Perusahaan'),
        ('gd_income', 'Penghasilan Bulanan'),
        ('gd_burden_count', 'Jumlah Tanggungan'), #DH
    ]
    col = 1
    row = 5
    for _, label in labels:
        ws.cell(row=row, column=col).value = label
        col += 1
    
    row += 1
    for row_data in data:
        col = 1
        for key, _ in labels:
            ws.cell(row=row, column=col).value = getattr(row_data, key)
            col += 1
        row += 1
        
    with NamedTemporaryFile(suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        file_content = tmp.read()
    return file_content